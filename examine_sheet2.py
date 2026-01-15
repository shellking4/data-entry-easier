#!/usr/bin/env python3
"""
Script to find the exact column locations in Excel
"""
import openpyxl

def find_columns_in_excel(excel_path):
    """Find the exact columns in the Excel file"""
    print("=" * 80)
    print("SEARCHING FOR COLUMNS IN EXCEL")
    print("=" * 80)
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[wb.sheetnames[1]]  # Second sheet
    
    print(f"Sheet: {ws.title}")
    print(f"Max row: {ws.max_row}, Max column: {ws.max_column}\n")
    
    # Search for columns in first 50 rows
    target_keywords = {
        "Description commerciale": None,
        "Quantité *": None,
        "Valeur incoterm par ligne *": None
    }
    
    print("Searching for target columns in first 50 rows...\n")
    
    for row in range(1, min(51, ws.max_row + 1)):
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row, col).value
            if cell_value:
                cell_str = str(cell_value).strip()
                
                # Check for each target
                for keyword in target_keywords.keys():
                    if keyword in cell_str:
                        if target_keywords[keyword] is None:
                            target_keywords[keyword] = (row, col)
                            print(f"✓ Found '{keyword}' at Row {row}, Column {col}")
                            print(f"  Full text: {cell_str[:100]}...")
    
    print("\n" + "=" * 80)
    print("SUMMARY OF FOUND COLUMNS:")
    print("=" * 80)
    for keyword, location in target_keywords.items():
        if location:
            print(f"✓ {keyword}: Row {location[0]}, Column {location[1]}")
        else:
            print(f"✗ {keyword}: NOT FOUND")
    
    # If we found the columns, show some sample data
    print("\n" + "=" * 80)
    print("SAMPLE DATA FROM FOUND COLUMNS:")
    print("=" * 80)
    
    if all(loc is not None for loc in target_keywords.values()):
        # Assume the header is on the same row
        header_row = list(target_keywords.values())[0][0]
        
        # Print column headers
        print(f"\nHeaders from row {header_row}:")
        for col in range(1, min(20, ws.max_column + 1)):
            val = ws.cell(header_row, col).value
            if val:
                print(f"  Col {col}: {str(val)[:60]}")
        
        # Print sample data rows
        print(f"\nSample data rows (starting from row {header_row + 1}):")
        for row in range(header_row + 1, min(header_row + 6, ws.max_row + 1)):
            print(f"\nRow {row}:")
            for keyword, (h_row, h_col) in target_keywords.items():
                val = ws.cell(row, h_col).value
                print(f"  {keyword}: {val}")

if __name__ == "__main__":
    excel_path = "IDI VIDE.xlsx"
    find_columns_in_excel(excel_path)
