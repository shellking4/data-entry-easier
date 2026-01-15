#!/usr/bin/env python3
"""
Script to verify the populated Excel file
"""
import openpyxl

def verify_excel(excel_path):
    """Verify the populated Excel file"""
    print("=" * 80)
    print(f"VERIFYING {excel_path}")
    print("=" * 80)
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb[wb.sheetnames[1]]  # Second sheet
        
        print(f"Sheet: {ws.title}")
        print(f"Max row: {ws.max_row}\n")
        
        # Target columns
        col_desc = 3
        col_valeur = 4
        col_qty = 5
        
        start_row = 6
        
        # Check first 5 rows of data
        print("Checking first 5 rows of populated data:")
        print("-" * 80)
        print(f"{'Row':<5} | {'Description':<40} | {'Valeur':<10} | {'Quantité':<10}")
        print("-" * 80)
        
        data_count = 0
        for row in range(start_row, min(start_row + 25, ws.max_row + 1)):
            desc = ws.cell(row, col_desc).value
            valeur = ws.cell(row, col_valeur).value
            qty = ws.cell(row, col_qty).value
            
            if desc or valeur or qty:
                print(f"{row:<5} | {str(desc)[:40]:<40} | {str(valeur):<10} | {str(qty):<10}")
                data_count += 1
        
        print("-" * 80)
        print(f"\nTotal rows with data found: {data_count}")
        
        if data_count > 0:
            print("\nVERIFICATION SUCCESSFUL: Data has been populated.")
        else:
            print("\nVERIFICATION FAILED: No data found in target rows.")
            
    except Exception as e:
        print(f"\nVERIFICATION FAILED: Error reading file - {e}")

if __name__ == "__main__":
    excel_path = "IDI_FILLED.xlsx"
    verify_excel(excel_path)