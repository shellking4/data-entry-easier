import openpyxl
import shutil
import os

def inspect_cell_e2(file_path):
    print(f"\nInspecting {file_path} - Cell E2...")
    wb = openpyxl.load_workbook(file_path)
    ws = wb[wb.sheetnames[1]]
    
    cell = ws['E2']
    print(f"Value: {cell.value}")
    if cell.hyperlink:
        print(f"Hyperlink Type: {type(cell.hyperlink)}")
        print(f"Target: {cell.hyperlink.target}")
        print(f"Location: {cell.hyperlink.location}")
        print(f"Display: {cell.hyperlink.display}")
        print(f"Tooltip: {cell.hyperlink.tooltip}")
    else:
        print("No hyperlink on E2")

# Test 1: Inspect original
inspect_cell_e2("IDI VIDE.xlsx")

# Test 2: Save without changes
print("\n--- TEST: Save without changes ---")
shutil.copy("IDI VIDE.xlsx", "IDI_TEST_SAVE.xlsx")
wb = openpyxl.load_workbook("IDI_TEST_SAVE.xlsx")
wb.save("IDI_TEST_SAVE.xlsx")
inspect_cell_e2("IDI_TEST_SAVE.xlsx")

# Test 3: Inspect filled
inspect_cell_e2("IDI_FILLED.xlsx")
