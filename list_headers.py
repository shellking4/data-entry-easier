import openpyxl

def list_headers(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[wb.sheetnames[1]]
    
    print(f"Sheet: {ws.title}")
    header_row = 5
    
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(header_row, col).value
        if val:
            headers.append((col, str(val).strip()))
            print(f"Col {col}: {val}")
            
    return headers

list_headers("IDI VIDE.xlsx")
