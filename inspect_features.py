import openpyxl

def inspect_features(file_path):
    print(f"Inspecting {file_path}...")
    wb = openpyxl.load_workbook(file_path)
    ws = wb[wb.sheetnames[1]]  # Marchandises importées
    
    print(f"Sheet: {ws.title}")
    
    # Check for Hyperlinks
    print("\nHyperlinks:")
    hyperlinks_found = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.hyperlink:
                print(f"  Cell {cell.coordinate}: {cell.hyperlink.target}")
                hyperlinks_found += 1
    if hyperlinks_found == 0:
        print("  No cell hyperlinks found.")
        
    # Check for Images/Drawings
    print("\nImages/Drawings:")
    if hasattr(ws, '_images') and ws._images:
        print(f"  Found {len(ws._images)} images.")
    else:
        print("  No images found (via _images).")
        
    # Check for legacy drawing objects (sometimes used for buttons)
    if hasattr(ws, 'legacy_drawing') and ws.legacy_drawing:
        print(f"  Found legacy drawing: {ws.legacy_drawing}")
    else:
        print("  No legacy drawing found.")

print("--- ORIGINAL TEMPLATE ---")
inspect_features("IDI VIDE.xlsx")

print("\n--- GENERATED FILE ---")
try:
    inspect_features("IDI_FILLED.xlsx")
except FileNotFoundError:
    print("IDI_FILLED.xlsx not found. Please generate it first.")
