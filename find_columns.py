#!/usr/bin/env python3
"""
Script to find all columns including Valeur incoterm
"""
import openpyxl

def find_all_columns(excel_path):
    """Find all target columns in the Excel file"""
    print("=" * 80)
    print("SEARCHING FOR ALL COLUMNS IN EXCEL")
    print("=" * 80)
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[wb.sheetnames[1]]  # Second sheet
    
    print(f"Sheet: {ws.title}")
    print(f"Max row: {ws.max_row}, Max column: {ws.max_column}\n")
    
    # Search for columns in first 50 rows
    print("Searching row 5 for all columns...\n")
    
    row = 5  # Based on previous search
    for col in range(1, min(45, ws.max_column + 1)):
        cell_value = ws.cell(row, col).value
        if cell_value:
            cell_str = str(cell_value).strip()
            print(f"Column {col}: {cell_str[:80]}")
    
    # Now search for "Valeur" keyword
    print("\n" + "=" * 80)
    print("Searching for 'Valeur' keyword in first 50 rows:")
    print("=" * 80)
    
    for row in range(1, min(51, ws.max_row + 1)):
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row, col).value
            if cell_value and 'Valeur' in str(cell_value):
                print(f"Row {row}, Column {col}: {str(cell_value)[:100]}")

if __name__ == "__main__":
    excel_path = "IDI VIDE.xlsx"
    find_all_columns(excel_path)
