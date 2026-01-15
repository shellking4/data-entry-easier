import streamlit as st
import pdfplumber
import openpyxl
import re
import zipfile
import shutil
import os
import xml.etree.ElementTree as ET
from io import BytesIO
import pandas as pd
import tempfile
import ocr_utils

# Namespaces
NS = {'x': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
ET.register_namespace('', NS['x'])

def get_excel_headers(template_path):
    """Extract headers from the Excel template (Row 5)"""
    wb = openpyxl.load_workbook(template_path)
    if len(wb.sheetnames) > 1:
        ws = wb[wb.sheetnames[1]]
    else:
        ws = wb.active
        
    headers = []
    header_row = 5
    for col in range(1, ws.max_column + 1):
        val = ws.cell(header_row, col).value
        if val:
            # Clean header: remove newlines, extra spaces
            clean_val = str(val).replace('\n', ' ').strip()
            headers.append((col, clean_val))
    return headers

def get_pdf_headers(pdf_file):
    """Extract headers from the first table in the PDF"""
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    row_clean = [str(cell).strip() for cell in row if cell]
                    if len(row_clean) > 2:
                        return [str(cell).strip() if cell else f"Col_{i}" for i, cell in enumerate(row)]
    return []

def get_input_excel_headers(excel_file):
    """Extract headers from the uploaded Excel file"""
    try:
        df = pd.read_excel(excel_file)
        # Convert to string and strip
        return [str(c).strip() for c in df.columns]
    except Exception as e:
        st.error(f"Error reading Excel file: {e}")
        return []

def extract_input_excel_data(excel_file, selected_headers):
    """Extract data from Excel file"""
    try:
        df = pd.read_excel(excel_file)
        
        # Filter columns
        # We need to map selected headers to actual columns
        # But selected_headers ARE the actual columns (or close to it)
        
        data = []
        for _, row in df.iterrows():
            row_data = {}
            for h in selected_headers:
                if h in df.columns:
                    val = row[h]
                    if pd.notna(val):
                        row_data[h] = val
            
            if row_data:
                data.append(row_data)
                
        return data
    except Exception as e:
        st.error(f"Error extracting data from Excel: {e}")
        return []

def extract_pdf_data(pdf_file, selected_pdf_headers):
    """Extract data from PDF file object"""
    data = []
    
    # We need to find a table that contains the selected headers
    # If no headers selected, we can't find the table easily.
    if not selected_pdf_headers:
        return []

    # Store the column mapping once found to use for subsequent pages
    global_col_indices = None
    
    with pdfplumber.open(pdf_file) as pdf:
        for i, page in enumerate(pdf.pages):
            tables = page.extract_tables()
            
            for table in tables:
                header_row_idx = -1
                headers = []
                
                # Try to find header row in this table
                for idx, row in enumerate(table):
                    row_values = [str(cell).strip() for cell in row if cell]
                    matches = sum(1 for h in selected_pdf_headers if h in row_values)
                    if matches > 0:
                        header_row_idx = idx
                        headers = [str(cell).strip() if cell else f"Col_{c_i}" for c_i, cell in enumerate(row)]
                        break
                
                # If headers found, update global mapping
                if header_row_idx != -1:
                    # Map column names to indices
                    global_col_indices = {h: i for i, h in enumerate(headers)}
                    
                    # Identify number column for filtering
                    no_col_idx = -1
                    for h, idx in global_col_indices.items():
                        if h.lower() in ['no', 'no.', 'item', '#', 'n°', 'pos']:
                            no_col_idx = idx
                            break
                    
                    # Process rows after header
                    for row in table[header_row_idx+1:]:
                        if not row or all(cell is None or cell == "" for cell in row):
                            continue
                        
                        # Filter by number column if it exists
                        if no_col_idx != -1 and no_col_idx < len(row):
                            val = row[no_col_idx]
                            # Check if value is numeric (allow digits, maybe ending with dot)
                            if not val:
                                continue
                            val_str = str(val).strip()
                            if not val_str or not val_str.replace('.', '').isdigit():
                                continue
                            
                        row_data = {}
                        for h, idx in global_col_indices.items():
                            if idx < len(row):
                                row_data[h] = row[idx]
                        
                        if any(row_data.values()):
                            data.append(row_data)
                            
                # If no headers found, but we have a global mapping, assume continuation
                elif global_col_indices is not None:
                    # We assume the table structure is similar (continuation)
                    
                    # Re-identify number column from global mapping (indices are same)
                    no_col_idx = -1
                    for h, idx in global_col_indices.items():
                        if h.lower() in ['no', 'no.', 'item', '#', 'n°', 'pos']:
                            no_col_idx = idx
                            break

                    for row in table:
                        if not row or all(cell is None or cell == "" for cell in row):
                            continue
                            
                        # Filter by number column if it exists
                        if no_col_idx != -1 and no_col_idx < len(row):
                            val = row[no_col_idx]
                            if not val:
                                continue
                            val_str = str(val).strip()
                            if not val_str or not val_str.replace('.', '').isdigit():
                                continue
                            
                        row_data = {}
                        for h, idx in global_col_indices.items():
                            if idx < len(row):
                                row_data[h] = row[idx]
                        
                        if any(row_data.values()):
                            data.append(row_data)
    
    return data

def clean_number(value):
    """Clean numeric values"""
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return value
    
    val_str = str(value).strip()
    val_str = val_str.replace(',', '.')
    val_str = re.sub(r'[^\d.]', '', val_str)
    
    try:
        return float(val_str)
    except ValueError:
        return 0

def get_col_letter(col_idx):
    """Convert 1-based column index to letter"""
    string = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        string = chr(65 + remainder) + string
    return string

def update_cell(row, row_idx, col_idx, value, val_type, clear_style=False):
    """Update or create a cell in the row"""
    col_letter = get_col_letter(col_idx)
    cell_ref = f"{col_letter}{row_idx}"
    
    # Find existing cell
    cell = None
    for c in row.findall('x:c', NS):
        if c.get('r') == cell_ref:
            cell = c
            break
            
    if cell is None:
        cell = ET.Element(f"{{{NS['x']}}}c", {'r': cell_ref})
        row.append(cell)
    
    # Clear style if requested (forces default alignment, usually left for text)
    if clear_style and 's' in cell.attrib:
        del cell.attrib['s']
    
    # Clear children
    for child in list(cell):
        cell.remove(child)
        
    # Set value
    if val_type == 'str':
        cell.set('t', 'inlineStr')
        is_elem = ET.SubElement(cell, f"{{{NS['x']}}}is")
        t_elem = ET.SubElement(is_elem, f"{{{NS['x']}}}t")
        t_elem.text = str(value)
    else:
        # Numeric
        if 't' in cell.attrib:
            del cell.attrib['t']
        v_elem = ET.SubElement(cell, f"{{{NS['x']}}}v")
        v_elem.text = str(value)

def populate_excel(data, template_path, mapping, excel_headers):
    """Populate Excel file using direct XML patching
    mapping: dict {excel_col_idx: [pdf_col_names]}
    excel_headers: list of (col_idx, col_name) tuples
    """
    
    # Create a temp file for the output
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        output_path = tmp.name
    
    shutil.copy(template_path, output_path)
    
    with zipfile.ZipFile(output_path, 'r') as zin:
        sheet_xml = zin.read('xl/worksheets/sheet2.xml')
    
    root = ET.fromstring(sheet_xml)
    sheetData = root.find('x:sheetData', NS)
    
    if sheetData is None:
        st.error("Error: sheetData not found in template")
        return None
        
    start_row = 6
    rows = {int(r.get('r')): r for r in sheetData.findall('x:row', NS)}
    
    for i, item in enumerate(data):
        row_idx = start_row + i
        
        if row_idx in rows:
            row = rows[row_idx]
        else:
            row = ET.Element(f"{{{NS['x']}}}row", {'r': str(row_idx)})
            sheetData.append(row)
            rows[row_idx] = row

        # Apply mapping
        for excel_col_idx, pdf_cols in mapping.items():
            if not pdf_cols:
                continue
                
            # Concatenate values
            parts = []
            for pdf_col in pdf_cols:
                val = item.get(pdf_col)
                if val:
                    parts.append(str(val).strip())
            
            final_val = " ".join(parts)
            
            # Check if this is the description column and uppercase it
            col_name = next((name for idx, name in excel_headers if idx == excel_col_idx), "")
            is_description = "description" in col_name.lower()
            if is_description:
                final_val = final_val.upper()
            
            # Determine type
            # Heuristic: if column name implies number, try to clean
            # For now, let's try to convert to float. If it works, use number.
            # Unless it's a code that looks like a number but should be string?
            # Excel handles numbers best as numbers.
            
            val_type = 'str'
            num_val = clean_number(final_val)
            
            if final_val and re.match(r'^-?\d+(\.\d+)?$', final_val.replace(',', '.')):
                 final_val = num_val
                 val_type = 'num'
            
            update_cell(row, row_idx, excel_col_idx, final_val, val_type, clear_style=is_description)
        
    temp_zip = output_path + ".tmp"
    with zipfile.ZipFile(template_path, 'r') as zin:
        with zipfile.ZipFile(temp_zip, 'w') as zout:
            for item in zin.infolist():
                if item.filename == 'xl/worksheets/sheet2.xml':
                    zout.writestr(item, ET.tostring(root, encoding='UTF-8', xml_declaration=True))
                else:
                    zout.writestr(item, zin.read(item.filename))
    
    shutil.move(temp_zip, output_path)
    
    with open(output_path, 'rb') as f:
        output = BytesIO(f.read())
        
    os.unlink(output_path)
    return output

def main():
    st.set_page_config(page_title="PDF to Excel Converter", layout="wide")
    
    st.title("📄 PDF to Excel Converter")
    st.markdown("""
    Upload your PDF invoice to automatically fill the data into the IDI template.
    """)
    
    template_path = "IDI VIDE.xlsx"
    try:
        open(template_path, 'rb')
    except FileNotFoundError:
        st.error(f"Template file '{template_path}' not found in the directory!")
        return
    
    uploaded_file = st.file_uploader("Upload Invoice (PDF or Excel)", type=["pdf", "xlsx"])
        
    if uploaded_file:
        file_type = uploaded_file.name.split('.')[-1].lower()
        
        input_headers = []
        processing_file_path = None
        is_pdf = False
        
        if file_type == 'pdf':
            is_pdf = True
            # Save uploaded file to temp file for processing
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_pdf:
                tmp_pdf.write(uploaded_file.getvalue())
                tmp_pdf_path = tmp_pdf.name
                
            try:
                # Check if OCR is needed
                if ocr_utils.needs_ocr(tmp_pdf_path):
                    st.warning("Scanned document detected. Performing OCR (this may take a while)...")
                    with st.spinner("Running OCR..."):
                        # Convert to searchable PDF
                        searchable_pdf_path = ocr_utils.convert_to_searchable_pdf(tmp_pdf_path)
                        # Use the new PDF for processing
                        processing_file_path = searchable_pdf_path
                else:
                    processing_file_path = tmp_pdf_path

                input_headers = get_pdf_headers(processing_file_path)
            except Exception as e:
                st.error(f"Error processing PDF: {e}")
                
        elif file_type == 'xlsx':
            input_headers = get_input_excel_headers(uploaded_file)
            # For Excel, we can just use the uploaded_file object directly with pandas, 
            # but for consistency we might want to just pass it.
            # However, extract_input_excel_data takes the file object.
            pass
            
        excel_headers = get_excel_headers(template_path)
        
        if input_headers and excel_headers:
            st.subheader("Column Mapping")
            st.info(f"Map the {file_type.upper()} columns to the Excel fields.")
            
            mapping = {}
            selected_input_headers = set()
            
            # Create 3 columns for layout
            cols = st.columns(3)
            
            for i, (col_idx, col_name) in enumerate(excel_headers):
                with cols[i % 3]:
                    # Default selection logic
                    default = []
                    col_name_lower = col_name.lower()
                    
                    for ph in input_headers:
                        ph_lower = ph.lower()
                        # Heuristics for defaults
                        if "description" in col_name_lower and ("model" in ph_lower or "material" in ph_lower):
                            default.append(ph)
                        elif "quantité" in col_name_lower and "qty" in ph_lower:
                            default.append(ph)
                        elif "valeur" in col_name_lower and "amount" in ph_lower:
                            default.append(ph)
                    
                    selection = st.multiselect(
                        f"{col_name}",
                        options=input_headers,
                        default=default,
                        key=f"map_{col_idx}"
                    )
                    mapping[col_idx] = selection
                    selected_input_headers.update(selection)
            
            if st.button("Process File", type="primary"):
                with st.spinner("Processing..."):
                    try:
                        data = []
                        if is_pdf:
                            data = extract_pdf_data(processing_file_path, list(selected_input_headers))
                        else:
                            # Excel
                            # Reset pointer to beginning of file if it was read before
                            uploaded_file.seek(0)
                            data = extract_input_excel_data(uploaded_file, list(selected_input_headers))
                            
                        st.success(f"Extracted {len(data)} items from {file_type.upper()}.")
                        
                        if data:
                            processed_excel = populate_excel(data, template_path, mapping, excel_headers)
                            
                            st.subheader("Preview of Data to be Written")
                            
                            # Create preview dataframe with mapped values
                            preview_rows = []
                            # Create a map of col_idx -> col_name for easy lookup
                            col_name_map = {idx: name for idx, name in excel_headers}
                            
                            for item in data:
                                row_data = {}
                                for col_idx, input_cols in mapping.items():
                                    if not input_cols:
                                        continue
                                    
                                    # Replicate the concatenation logic
                                    parts = []
                                    for input_col in input_cols:
                                        val = item.get(input_col)
                                        if val:
                                            parts.append(str(val).strip())
                                    
                                    final_val = " ".join(parts)
                                    col_name = col_name_map.get(col_idx, f"Col {col_idx}")
                                    row_data[col_name] = final_val
                                
                                if row_data:
                                    preview_rows.append(row_data)
                            
                            if preview_rows:
                                st.dataframe(pd.DataFrame(preview_rows))
                            else:
                                st.info("No data mapped yet.")
                            
                            st.download_button(
                                label="📥 Download Filled Excel",
                                data=processed_excel,
                                file_name="IDI_FILLED.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        else:
                            st.warning(f"No data found in the {file_type.upper()} matching the selected columns.")
                            
                    except Exception as e:
                        st.error(f"An error occurred: {str(e)}")
        else:
            if not input_headers:
                st.warning(f"Could not detect headers in the {file_type.upper()}.")
            if not excel_headers:
                st.error("Could not read headers from Excel template.")
        
        # Cleanup
        if is_pdf:
            if 'tmp_pdf_path' in locals() and os.path.exists(tmp_pdf_path):
                os.unlink(tmp_pdf_path)
            if 'processing_file_path' in locals() and processing_file_path != tmp_pdf_path and os.path.exists(processing_file_path):
                os.unlink(processing_file_path)
                    
    elif not uploaded_file:
        st.info("Please upload a PDF or Excel file to start.")

if __name__ == "__main__":
    main()
