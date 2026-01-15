#!/usr/bin/env python3
"""
Script to examine the structure of PDF and Excel files
"""
import pdfplumber
import openpyxl
import pandas as pd

def examine_pdf(pdf_path):
    """Examine PDF structure and extract sample data"""
    print("=" * 80)
    print("EXAMINING PDF FILE")
    print("=" * 80)
    
    with pdfplumber.open(pdf_path) as pdf:
        print(f"Total pages: {len(pdf.pages)}\n")
        
        # Examine first page
        first_page = pdf.pages[0]
        print("First page text (first 500 chars):")
        print(first_page.extract_text()[:500])
        print("\n")
        
        # Try to extract tables
        print("Attempting to extract tables from first page:")
        tables = first_page.extract_tables()
        if tables:
            print(f"Found {len(tables)} table(s)")
            for i, table in enumerate(tables):
                print(f"\nTable {i+1}:")
                print(f"Rows: {len(table)}, Columns: {len(table[0]) if table else 0}")
                if table:
                    # Print first few rows
                    print("First 3 rows:")
                    for row_idx, row in enumerate(table[:3]):
                        print(f"Row {row_idx}: {row}")
        else:
            print("No tables found using default extraction")
            
        # Try extracting all text from first page
        print("\n" + "=" * 80)
        print("Full text from first page:")
        print("=" * 80)
        print(first_page.extract_text())

def examine_excel(excel_path):
    """Examine Excel structure"""
    print("\n" + "=" * 80)
    print("EXAMINING EXCEL FILE")
    print("=" * 80)
    
    # Load workbook
    wb = openpyxl.load_workbook(excel_path)
    print(f"Sheet names: {wb.sheetnames}\n")
    
    # Examine first sheet
    ws = wb.active
    print(f"Active sheet: {ws.title}")
    print(f"Max row: {ws.max_row}, Max column: {ws.max_column}\n")
    
    # Print headers (first row)
    print("Headers (first row):")
    headers = []
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(1, col).value
        headers.append(cell_value)
        print(f"Column {col}: {cell_value}")
    
    # Print first few data rows
    print("\nFirst 3 data rows:")
    for row in range(2, min(5, ws.max_row + 1)):
        print(f"\nRow {row}:")
        for col in range(1, min(10, ws.max_column + 1)):  # Limit to first 10 columns
            cell_value = ws.cell(row, col).value
            print(f"  Column {col} ({headers[col-1]}): {cell_value}")
    
    # Look for target columns
    print("\n" + "=" * 80)
    print("Looking for target columns:")
    print("=" * 80)
    target_columns = [
        "Description commerciale des marchandises",
        "Quantité *",
        "Valeur incoterm par ligne *"
    ]
    
    for target in target_columns:
        found = False
        for col_idx, header in enumerate(headers, 1):
            if header and target.lower() in str(header).lower():
                print(f"✓ Found '{target}' at column {col_idx}: '{header}'")
                found = True
                break
        if not found:
            print(f"✗ Column '{target}' not found")

if __name__ == "__main__":
    pdf_path = "fa.pdf"
    excel_path = "IDI VIDE.xlsx"
    
    examine_pdf(pdf_path)
    examine_excel(excel_path)
