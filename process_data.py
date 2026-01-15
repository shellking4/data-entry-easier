import pdfplumber
import openpyxl
import pandas as pd
import re

def extract_pdf_data(pdf_path):
    """Extract data from PDF file"""
    print(f"Extracting data from {pdf_path}...")
    
    data = []
    
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages):
            print(f"Processing page {i+1}...")
            
            # Extract tables
            tables = page.extract_tables()
            
            for table in tables:
                # Check if this is the data table by looking for headers
                # Based on analysis: ['no', 'product models', 'QTY', 'Price', 'AMOUNT', 'CTN', 'PHOTOS', 'Materials']
                header_row_idx = -1
                for idx, row in enumerate(table):
                    # Clean row values
                    row_clean = [str(cell).strip().lower() if cell else "" for cell in row]
                    if "product models" in row_clean and "qty" in row_clean:
                        header_row_idx = idx
                        print(f"Found header row at index {idx}")
                        break
                
                if header_row_idx != -1:
                    # Map columns
                    headers = [str(cell).strip().lower() if cell else "" for cell in table[header_row_idx]]
                    try:
                        model_idx = -1
                        qty_idx = -1
                        amount_idx = -1
                        material_idx = -1
                        
                        # Find indices (allowing for some fuzzy matching or exact known positions)
                        for idx, h in enumerate(headers):
                            if "product models" in h: model_idx = idx
                            elif "qty" in h: qty_idx = idx
                            elif "amount" in h: amount_idx = idx
                            elif "materials" in h: material_idx = idx
                        
                        # If headers are not found by name, try fixed indices based on visual inspection
                        # no(0) product models(1) QTY(2) Price(3) AMOUNT(4) CTN(5) PHOTOS(6) Materials(7)
                        if model_idx == -1: model_idx = 1
                        if qty_idx == -1: qty_idx = 2
                        if amount_idx == -1: amount_idx = 4
                        if material_idx == -1: material_idx = 7
                        
                        print(f"Column Mapping: Model={model_idx}, Qty={qty_idx}, Amount={amount_idx}, Material={material_idx}")
                        
                        # Extract data rows
                        for row in table[header_row_idx+1:]:
                            # Skip empty rows or summary rows
                            if not row or all(cell is None or cell == "" for cell in row):
                                continue
                                
                            # Check if row has enough columns
                            if len(row) <= max(model_idx, qty_idx, amount_idx, material_idx):
                                continue
                                
                            model = row[model_idx]
                            qty = row[qty_idx]
                            amount = row[amount_idx]
                            material = row[material_idx]
                            
                            # Skip if model is empty (likely not a data row)
                            if not model:
                                continue
                                
                            data.append({
                                "model": str(model).strip(),
                                "qty": qty,
                                "amount": amount,
                                "material": str(material).strip() if material else ""
                            })
                            
                    except Exception as e:
                        print(f"Error processing table: {e}")
    
    print(f"Extracted {len(data)} items.")
    return data

def clean_number(value):
    """Clean numeric values"""
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return value
    
    # Remove currency symbols, spaces, etc.
    val_str = str(value).strip()
    # Replace comma with dot if needed
    val_str = val_str.replace(',', '.')
    # Remove non-numeric chars except dot
    val_str = re.sub(r'[^\d.]', '', val_str)
    
    try:
        return float(val_str)
    except ValueError:
        return 0

def populate_excel(data, input_excel, output_excel):
    """Populate Excel file with extracted data"""
    print(f"Opening Excel file {input_excel}...")
    
    wb = openpyxl.load_workbook(input_excel)
    
    # Select the second sheet "Marchandises importées"
    if len(wb.sheetnames) > 1:
        ws = wb[wb.sheetnames[1]]
    else:
        ws = wb.active
        
    print(f"Working on sheet: {ws.title}")
    
    # Target columns (1-based index)
    # Description commerciale des marchandises *: Column 3
    # Valeur incoterm par ligne *: Column 4
    # Quantité *: Column 5
    col_desc = 3
    col_valeur = 4
    col_qty = 5
    
    # Start row for data entry (Row 6 based on analysis)
    start_row = 6
    
    print(f"Writing data starting at row {start_row}...")
    
    for i, item in enumerate(data):
        row_idx = start_row + i
        
        # Description = model + " " + material
        description = f"{item['model']} {item['material']}".strip()
        
        # Clean numbers
        qty = clean_number(item['qty'])
        amount = clean_number(item['amount'])
        
        # Write to cells
        ws.cell(row=row_idx, column=col_desc).value = description
        ws.cell(row=row_idx, column=col_valeur).value = amount
        ws.cell(row=row_idx, column=col_qty).value = qty
        
    print(f"Saving to {output_excel}...")
    wb.save(output_excel)
    print("Done!")

if __name__ == "__main__":
    pdf_file = "fa.pdf"
    input_excel = "IDI VIDE.xlsx"
    output_excel = "IDI_FILLED.xlsx"
    
    extracted_data = extract_pdf_data(pdf_file)
    
    if extracted_data:
        # Show sample
        print("\nSample extracted data:")
        for item in extracted_data[:3]:
            print(item)
            
        populate_excel(extracted_data, input_excel, output_excel)
    else:
        print("No data extracted from PDF!")
