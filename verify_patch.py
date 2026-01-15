import openpyxl

def verify_patched(file_path):
    print(f"Verifying {file_path}...")
    wb = openpyxl.load_workbook(file_path)
    ws = wb[wb.sheetnames[1]]
    
    print(f"Row 6, Col 3: {ws.cell(6, 3).value}")
    print(f"Row 6, Col 4: {ws.cell(6, 4).value}")
    print(f"Row 6, Col 5: {ws.cell(6, 5).value}")

verify_patched("IDI_PATCHED.xlsx")
